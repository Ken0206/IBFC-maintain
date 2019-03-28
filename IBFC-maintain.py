#!python
# -*- coding: utf8 -*-
# -*- coding: cp950 -*-　

import openpyxl, csv, glob
from openpyxl.styles import Font, Alignment, PatternFill

def importXlsx(sheetName, csvFileName):
    ws = wb.create_sheet(sheetName)
    ws.append(["Option", "Result"])
    a1 = ws['A1']
    b1 = ws['B1']
    font = Font(color="FFFFFF")
    alignment = Alignment(horizontal='center', vertical='center')
    fill = PatternFill("solid", fgColor="757171")
    a1.font = font
    b1.font = font
    a1.alignment = alignment
    b1.alignment = alignment
    a1.fill = fill
    b1.fill = fill
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 60
    f = open(csvFileName)
    reader = csv.reader(f, delimiter='|')
    for row in reader:
        ws.append(row)
    f.close()

xlsx = []
for file in glob.glob("*.txt"):
    xlsx.append(file[:1])
    xlsx = list(set(xlsx))
    xlsx.sort()
for i in xlsx:
    xlsxFileName = i+".xlsx"
    wb = openpyxl.Workbook()
    ii = str(i)+'-*.txt'
    for csvFileName in glob.glob(ii):
        sheetName = csvFileName[:len(csvFileName)-4]
        importXlsx(sheetName, csvFileName)
    wb.remove(wb['Sheet'])
    wb.save(xlsxFileName)


